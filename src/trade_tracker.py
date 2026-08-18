"""Excel journal for scheduled swing picks and one-month performance review."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.constants import DIRECTION_LONG
from src.data_fetcher import DataFetcher


PICK_HEADERS = [
    "Trade ID",
    "Scan Date",
    "Week Start",
    "Rank",
    "Symbol",
    "Direction",
    "Setup",
    "Conviction",
    "Quant Score",
    "AI Score",
    "Entry",
    "Stop Loss",
    "Target 1",
    "Target 2",
    "Quantity",
    "Risk Amount",
    "Status",
    "Trigger Date",
    "Target 1 Date",
    "Exit Date",
    "Exit Price",
    "Return %",
    "R Multiple",
    "Holding Sessions",
    "Last Updated",
    "Notes",
]

RUN_HEADERS = [
    "Run Date",
    "Started At",
    "Completed At",
    "Status",
    "Picks",
    "Report File",
    "Message",
]

ACTIVE_STATUSES = {"PENDING", "OPEN", "TARGET1_HIT"}


def _week_start(day: date) -> date:
    return day - timedelta(days=day.weekday())


class TradeTracker:
    def __init__(self, path: Path, max_hold_sessions: int = 10) -> None:
        self.path = path
        self.max_hold_sessions = max_hold_sessions
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        if self.path.exists():
            return
        wb = Workbook()
        picks = wb.active
        picks.title = "Picks"
        picks.append(PICK_HEADERS)
        runs = wb.create_sheet("Runs")
        runs.append(RUN_HEADERS)
        wb.create_sheet("Summary")
        self._style_sheet(picks)
        self._style_sheet(runs)
        self._write_summary(wb)
        wb.save(self.path)

    @staticmethod
    def _style_sheet(ws) -> None:
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, header in enumerate(ws[1], 1):
            ws.column_dimensions[get_column_letter(idx)].width = min(
                max(len(str(header.value)) + 3, 12), 26
            )

    def already_ran_successfully(self, day: date) -> bool:
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb["Runs"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                run_day = _as_date(row[0])
                if run_day == day and str(row[3]).upper() == "SUCCESS":
                    return True
            return False
        finally:
            wb.close()

    def log_run(
        self,
        *,
        run_date: date,
        started_at: datetime,
        status: str,
        picks: int,
        report_file: str,
        message: str,
    ) -> None:
        wb = load_workbook(self.path)
        ws = wb["Runs"]
        ws.append(
            [
                run_date.isoformat(),
                started_at.isoformat(timespec="seconds"),
                datetime.now().isoformat(timespec="seconds"),
                status,
                picks,
                report_file,
                message,
            ]
        )
        self._style_sheet(ws)
        self._write_summary(wb)
        wb.save(self.path)

    def append_picks(self, result: dict[str, Any]) -> int:
        picks = result.get("top_trades") or []
        if not picks:
            return 0

        scan_day = _as_date(result.get("session_date")) or date.today()
        week = _week_start(scan_day)
        ai_verdict = result.get("ai_verdict")
        ai_scores = {
            pick.symbol: pick.ai_score
            for pick in getattr(ai_verdict, "picks", [])
        }

        wb = load_workbook(self.path)
        ws = wb["Picks"]
        existing = {
            (str(row[2].value), str(row[4].value).upper())
            for row in ws.iter_rows(min_row=2)
        }
        added = 0

        for rank, candidate in enumerate(picks, 1):
            a, p = candidate.analysis, candidate.plan
            key = (week.isoformat(), a.symbol.upper())
            if key in existing:
                continue
            trade_id = f"{week:%Y%m%d}-{rank}-{a.symbol.upper()}"
            ws.append(
                [
                    trade_id,
                    scan_day.isoformat(),
                    week.isoformat(),
                    candidate.week_rank or rank,
                    a.symbol.upper(),
                    p.direction,
                    a.setup,
                    a.conviction,
                    a.score,
                    ai_scores.get(a.symbol, ""),
                    p.entry,
                    p.stop_loss,
                    p.target_1,
                    p.target,
                    p.quantity,
                    p.risk_amount,
                    "PENDING",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    0,
                    scan_day.isoformat(),
                    f"{p.level_basis}; hold {p.hold_until}",
                ]
            )
            existing.add(key)
            added += 1

        self._style_sheet(ws)
        self._write_summary(wb)
        wb.save(self.path)
        return added

    def append_manual_records(
        self,
        records: list[dict[str, Any]],
        scan_date: date,
    ) -> int:
        """Backfill picks produced before Excel tracking was enabled."""
        if not records:
            return 0

        week = _week_start(scan_date)
        wb = load_workbook(self.path)
        ws = wb["Picks"]
        existing = {
            (str(row[2].value), str(row[4].value).upper())
            for row in ws.iter_rows(min_row=2)
        }
        added = 0

        for rank, record in enumerate(records, 1):
            symbol = str(record["symbol"]).upper()
            key = (week.isoformat(), symbol)
            if key in existing:
                continue
            ws.append(
                [
                    f"{week:%Y%m%d}-{rank}-{symbol}",
                    scan_date.isoformat(),
                    week.isoformat(),
                    record.get("rank", rank),
                    symbol,
                    record["direction"],
                    record["setup"],
                    record.get("conviction", ""),
                    record.get("quant_score", ""),
                    record.get("ai_score", ""),
                    record["entry"],
                    record["stop_loss"],
                    record["target_1"],
                    record["target_2"],
                    record["quantity"],
                    record.get("risk_amount", ""),
                    "PENDING",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    0,
                    scan_date.isoformat(),
                    record.get("notes", "Backfilled from manual scan"),
                ]
            )
            existing.add(key)
            added += 1

        self._style_sheet(ws)
        self._write_summary(wb)
        wb.save(self.path)
        return added

    def update_active_trades(self, fetcher: DataFetcher) -> int:
        wb = load_workbook(self.path)
        ws = wb["Picks"]
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
        updated = 0

        for row_idx in range(2, ws.max_row + 1):
            status = str(ws.cell(row_idx, header_map["Status"]).value or "")
            if status not in ACTIVE_STATUSES:
                continue

            symbol = str(ws.cell(row_idx, header_map["Symbol"]).value)
            scan_date = _as_date(ws.cell(row_idx, header_map["Scan Date"]).value)
            if not scan_date:
                continue

            daily = fetcher.fetch_daily(symbol, days=120)
            if daily.empty:
                continue

            values = {
                name: ws.cell(row_idx, col).value
                for name, col in header_map.items()
            }
            outcome = _evaluate_trade(
                daily=daily,
                direction=str(values["Direction"]),
                scan_date=scan_date,
                entry=float(values["Entry"]),
                stop=float(values["Stop Loss"]),
                target_1=float(values["Target 1"]),
                target_2=float(values["Target 2"]),
                max_hold_sessions=self.max_hold_sessions,
            )
            for name, value in outcome.items():
                if name in header_map:
                    ws.cell(row_idx, header_map[name], value)
            ws.cell(
                row_idx,
                header_map["Last Updated"],
                daily.index[-1].date().isoformat(),
            )
            updated += 1

        self._style_sheet(ws)
        self._write_summary(wb)
        wb.save(self.path)
        return updated

    def _write_summary(self, wb) -> None:
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("Summary")

        picks = wb["Picks"]
        headers = {cell.value: idx for idx, cell in enumerate(picks[1], 1)}
        rows = list(picks.iter_rows(min_row=2, values_only=True))
        statuses = [str(r[headers["Status"] - 1] or "") for r in rows]
        returns = [
            float(r[headers["Return %"] - 1])
            for r in rows
            if r[headers["Return %"] - 1] not in (None, "")
        ]
        r_values = [
            float(r[headers["R Multiple"] - 1])
            for r in rows
            if r[headers["R Multiple"] - 1] not in (None, "")
        ]
        closed = [
            s for s in statuses if s not in ACTIVE_STATUSES and s != ""
        ]
        wins = [r for r in returns if r > 0]

        metrics = [
            ("As of", datetime.now().isoformat(timespec="seconds")),
            ("Total picks", len(rows)),
            ("Pending / active", sum(s in ACTIVE_STATUSES for s in statuses)),
            ("Closed", len(closed)),
            ("Wins", len(wins)),
            ("Losses", sum(r < 0 for r in returns)),
            ("Win rate", round((len(wins) / len(closed) * 100), 2) if closed else 0),
            ("Average return %", round(sum(returns) / len(returns), 2) if returns else 0),
            ("Average R", round(sum(r_values) / len(r_values), 2) if r_values else 0),
            ("Cumulative trade return %", round(sum(returns), 2)),
        ]
        ws.append(["Performance Metric", "Value"])
        for item in metrics:
            ws.append(item)
        self._style_sheet(ws)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 22


def _evaluate_trade(
    *,
    daily,
    direction: str,
    scan_date: date,
    entry: float,
    stop: float,
    target_1: float,
    target_2: float,
    max_hold_sessions: int,
) -> dict[str, Any]:
    bars = daily[[ts.date() >= scan_date for ts in daily.index]]
    is_long = direction == DIRECTION_LONG
    triggered = False
    target1_hit = False
    trigger_date = ""
    target1_date = ""
    holding_sessions = 0

    for timestamp, bar in bars.iterrows():
        bar_date = timestamp.date()
        high = float(bar["high"])
        low = float(bar["low"])
        close = float(bar["close"])

        if not triggered:
            entry_hit = high >= entry if is_long else low <= entry
            if not entry_hit:
                continue
            triggered = True
            trigger_date = bar_date.isoformat()

        holding_sessions += 1
        active_stop = entry if target1_hit else stop
        stop_hit = low <= active_stop if is_long else high >= active_stop
        t1_hit_now = high >= target_1 if is_long else low <= target_1
        t2_hit_now = high >= target_2 if is_long else low <= target_2

        # Daily OHLC cannot tell which level came first. Use conservative ordering.
        if stop_hit:
            if target1_hit:
                return _closed_outcome(
                    status="BREAKEVEN_AFTER_T1",
                    trigger_date=trigger_date,
                    target1_date=target1_date,
                    exit_date=bar_date,
                    effective_exit=(target_1 + entry) / 2,
                    entry=entry,
                    stop=stop,
                    direction=direction,
                    holding_sessions=holding_sessions,
                    note="Second half stopped at cost after T1",
                )
            return _closed_outcome(
                status="STOPPED",
                trigger_date=trigger_date,
                target1_date="",
                exit_date=bar_date,
                effective_exit=stop,
                entry=entry,
                stop=stop,
                direction=direction,
                holding_sessions=holding_sessions,
                note=(
                    "Stop and target touched same daily bar; counted stop first"
                    if t1_hit_now or t2_hit_now
                    else "Structural stop hit"
                ),
            )

        if t2_hit_now:
            if not target1_hit:
                target1_hit = True
                target1_date = bar_date.isoformat()
            return _closed_outcome(
                status="TARGET2_HIT",
                trigger_date=trigger_date,
                target1_date=target1_date,
                exit_date=bar_date,
                effective_exit=(target_1 + target_2) / 2,
                entry=entry,
                stop=stop,
                direction=direction,
                holding_sessions=holding_sessions,
                note="Assumes 50% booked at T1 and 50% at T2",
            )

        if t1_hit_now and not target1_hit:
            target1_hit = True
            target1_date = bar_date.isoformat()

        if holding_sessions >= max_hold_sessions:
            effective_exit = (target_1 + close) / 2 if target1_hit else close
            return _closed_outcome(
                status="TIME_EXIT",
                trigger_date=trigger_date,
                target1_date=target1_date,
                exit_date=bar_date,
                effective_exit=effective_exit,
                entry=entry,
                stop=stop,
                direction=direction,
                holding_sessions=holding_sessions,
                note=f"Closed after {max_hold_sessions} sessions",
            )

    return {
        "Status": (
            "TARGET1_HIT" if target1_hit else "OPEN" if triggered else "PENDING"
        ),
        "Trigger Date": trigger_date,
        "Target 1 Date": target1_date,
        "Holding Sessions": holding_sessions,
    }


def _closed_outcome(
    *,
    status: str,
    trigger_date: str,
    target1_date: str,
    exit_date: date,
    effective_exit: float,
    entry: float,
    stop: float,
    direction: str,
    holding_sessions: int,
    note: str,
) -> dict[str, Any]:
    sign = 1 if direction == DIRECTION_LONG else -1
    pnl_per_share = (effective_exit - entry) * sign
    initial_risk = abs(entry - stop)
    return {
        "Status": status,
        "Trigger Date": trigger_date,
        "Target 1 Date": target1_date,
        "Exit Date": exit_date.isoformat(),
        "Exit Price": round(effective_exit, 2),
        "Return %": round((pnl_per_share / entry) * 100, 2),
        "R Multiple": round(pnl_per_share / initial_risk, 2)
        if initial_risk
        else 0,
        "Holding Sessions": holding_sessions,
        "Notes": note,
    }


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        try:
            return date.fromisoformat(str(value)[:10])
        except ValueError:
            return None
    return None
