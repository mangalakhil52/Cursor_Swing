"""Tests for Excel journal outcome tracking."""

from datetime import date

import pandas as pd
from openpyxl import load_workbook

from src.trade_tracker import TradeTracker, _evaluate_trade


def _bars(rows):
    index = pd.bdate_range("2026-08-03", periods=len(rows), tz="Asia/Kolkata")
    return pd.DataFrame(rows, index=index)


def test_creates_excel_workbook(tmp_path):
    path = tmp_path / "swing_performance.xlsx"
    TradeTracker(path)
    assert path.exists()
    wb = load_workbook(path, read_only=True)
    assert {"Picks", "Runs", "Summary"}.issubset(wb.sheetnames)
    wb.close()


def test_backfills_all_new_records_without_weekly_duplicates(tmp_path):
    path = tmp_path / "swing_performance.xlsx"
    tracker = TradeTracker(path)
    records = []
    for symbol in ("CUPID", "WELCORP", "RELIANCE"):
        records.append(
            {
                "symbol": symbol,
                "direction": "LONG",
                "setup": "Displacement Continuation",
                "entry": 244.61,
                "stop_loss": 232.98,
                "target_1": 256.24,
                "target_2": 262.91,
                "quantity": 86,
            }
        )

    assert tracker.append_manual_records(records, date(2026, 8, 5)) == 3
    assert tracker.append_manual_records(records, date(2026, 8, 5)) == 0

    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Picks"].iter_rows(min_row=2, values_only=True))
    wb.close()
    assert len(rows) == 3
    assert [row[4] for row in rows] == ["CUPID", "WELCORP", "RELIANCE"]


def test_long_tracks_partial_then_target2():
    daily = _bars(
        [
            {"open": 99, "high": 103, "low": 98, "close": 102, "volume": 1},
            {"open": 102, "high": 106, "low": 101, "close": 105, "volume": 1},
        ]
    )
    result = _evaluate_trade(
        daily=daily,
        direction="LONG",
        scan_date=date(2026, 8, 3),
        entry=100,
        stop=97,
        target_1=103,
        target_2=106,
        max_hold_sessions=10,
    )
    assert result["Status"] == "TARGET2_HIT"
    assert result["Exit Price"] == 104.5
    assert result["Return %"] == 4.5
    assert result["R Multiple"] == 1.5


def test_same_bar_stop_and_target_is_counted_conservatively():
    daily = _bars(
        [
            {"open": 100, "high": 104, "low": 96, "close": 101, "volume": 1},
        ]
    )
    result = _evaluate_trade(
        daily=daily,
        direction="LONG",
        scan_date=date(2026, 8, 3),
        entry=100,
        stop=97,
        target_1=103,
        target_2=106,
        max_hold_sessions=10,
    )
    assert result["Status"] == "STOPPED"
    assert result["R Multiple"] == -1


def test_untriggered_trade_stays_pending():
    daily = _bars(
        [
            {"open": 95, "high": 99, "low": 94, "close": 98, "volume": 1},
        ]
    )
    result = _evaluate_trade(
        daily=daily,
        direction="LONG",
        scan_date=date(2026, 8, 3),
        entry=100,
        stop=97,
        target_1=103,
        target_2=106,
        max_hold_sessions=10,
    )
    assert result["Status"] == "PENDING"
